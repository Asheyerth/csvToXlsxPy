import pandas as pd 
from tkinter.filedialog import askopenfilename
import tkinter as tk
from openpyxl import load_workbook
import openpyxl

#https://stackoverflow.com/questions/3579568/choosing-a-file-in-python-with-simple-dialog
filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file
print(filename)

#Leer el csv
tabla=pd.read_csv(filename, sep=',', encoding='utf-8')
print(tabla)

#Leer el nombre del archivo
index = filename.rfind('/')
name=filename[index+1:len(filename)]
nameFile=name+".xlsx"
print(nameFile)

#tabla.to_excel(nameFile, index=False)


wb = openpyxl.load_workbook(nameFile) 
res = len(wb.sheetnames)
print(res)
page = int(res)+1

#https://stackoverflow.com/questions/42370977/how-to-save-a-new-sheet-in-an-existing-excel-file-using-pandas
with pd.ExcelWriter(nameFile, engine='openpyxl', mode='a') as writer:  
    tabla.to_excel(writer, sheet_name=str(page), index=False)

